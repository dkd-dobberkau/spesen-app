#!/usr/bin/env python3
"""
Spesen CLI - Automatische Spesenabrechnung aus Belegen

Verwendung:
    python cli.py /pfad/zu/belegen --name "Max Mustermann" --monat "Dez 2025"
    python cli.py /pfad/zu/belegen --output spesen.xlsx --format excel
    python cli.py /pfad/zu/belegen --output spesen.pdf --format pdf
"""

import argparse
import os
import sys
import json
import sqlite3
import hashlib
from datetime import datetime
from pathlib import Path

# Imports aus der App
from dotenv import load_dotenv
load_dotenv()

# Datenbank-Pfad (gleiche wie Web-App, im data-Verzeichnis für Persistenz)
DATA_DIR = os.environ.get('DATA_DIR', os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data'))
os.makedirs(DATA_DIR, exist_ok=True)
DATABASE = os.path.join(DATA_DIR, 'spesen.db')

# Cache-Datei für extrahierte Beleg-Daten (im persistenten data-Verzeichnis)
CACHE_FILE = os.path.join(DATA_DIR, '.beleg_cache.json')


def get_file_hash(filepath):
    """Berechnet MD5-Hash einer Datei für Cache-Key"""
    hasher = hashlib.md5()
    with open(filepath, 'rb') as f:
        for chunk in iter(lambda: f.read(65536), b''):
            hasher.update(chunk)
    return hasher.hexdigest()


def load_cache():
    """Lädt den Cache aus der JSON-Datei"""
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            return {}
    return {}


def save_cache(cache):
    """Speichert den Cache in die JSON-Datei"""
    try:
        with open(CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except IOError as e:
        print(f"⚠️  Cache konnte nicht gespeichert werden: {e}")

from PIL import Image
import anthropic
import pytesseract
import base64
import io

# Optional: pdf2image
try:
    from pdf2image import convert_from_path
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False
    print("⚠️  pdf2image nicht installiert - PDF-Support deaktiviert")

# Excel/PDF Export
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

# Requests für Währungs-API (optional)
try:
    import requests
    REQUESTS_AVAILABLE = True
except ImportError:
    REQUESTS_AVAILABLE = False


# ============================================================================
# Währungsumrechnung
# ============================================================================

# Fallback-Wechselkurse zu EUR (Stand: Dezember 2024)
# Diese werden verwendet wenn keine API verfügbar ist
FALLBACK_EXCHANGE_RATES = {
    'EUR': 1.0,
    'USD': 0.95,      # 1 USD = 0.95 EUR
    'GBP': 1.17,      # 1 GBP = 1.17 EUR
    'CHF': 1.06,      # 1 CHF = 1.06 EUR
    'DKK': 0.134,     # 1 DKK = 0.134 EUR (Dänemark, Euro-gebunden)
    'SEK': 0.088,     # 1 SEK = 0.088 EUR (Schweden)
    'NOK': 0.085,     # 1 NOK = 0.085 EUR (Norwegen)
    'PLN': 0.23,      # 1 PLN = 0.23 EUR (Polen)
    'CZK': 0.040,     # 1 CZK = 0.040 EUR (Tschechien)
    'HUF': 0.0025,    # 1 HUF = 0.0025 EUR (Ungarn)
    'RON': 0.20,      # 1 RON = 0.20 EUR (Rumänien)
    'BGN': 0.51,      # 1 BGN = 0.51 EUR (Bulgarien, Euro-gebunden)
    'HRK': 0.133,     # 1 HRK = 0.133 EUR (Kroatien, historisch)
    'JPY': 0.0063,    # 1 JPY = 0.0063 EUR (Japan)
    'CNY': 0.13,      # 1 CNY = 0.13 EUR (China)
    'AUD': 0.61,      # 1 AUD = 0.61 EUR (Australien)
    'CAD': 0.68,      # 1 CAD = 0.68 EUR (Kanada)
}

# Cache für API-Wechselkurse (wird einmal pro Session geladen)
_exchange_rates_cache = None


def get_exchange_rates():
    """
    Holt aktuelle Wechselkurse von der EZB oder verwendet Fallback.
    Cached das Ergebnis für die gesamte Session.
    """
    global _exchange_rates_cache

    if _exchange_rates_cache is not None:
        return _exchange_rates_cache

    # Versuche EZB-Kurse zu laden (kostenlos, kein API-Key nötig)
    if REQUESTS_AVAILABLE:
        try:
            # EZB Exchange Rates API (XML)
            response = requests.get(
                'https://www.ecb.europa.eu/stats/eurofxref/eurofxref-daily.xml',
                timeout=5
            )
            if response.status_code == 200:
                import re
                rates = {'EUR': 1.0}
                # Parse XML einfach mit Regex (schneller als XML-Parser)
                for match in re.finditer(r"currency='(\w+)' rate='([\d.]+)'", response.text):
                    currency, rate = match.groups()
                    # EZB gibt EUR zu Fremdwährung, wir brauchen Fremdwährung zu EUR
                    rates[currency] = 1.0 / float(rate)

                if len(rates) > 5:  # Sanity check
                    print(f"💱 Aktuelle EZB-Wechselkurse geladen ({len(rates)} Währungen)")
                    _exchange_rates_cache = rates
                    return rates
        except Exception as e:
            print(f"⚠️  EZB-Kurse nicht verfügbar: {e}")

    # Fallback verwenden
    print("💱 Verwende Fallback-Wechselkurse")
    _exchange_rates_cache = FALLBACK_EXCHANGE_RATES.copy()
    return _exchange_rates_cache


def convert_to_eur(amount, currency):
    """
    Konvertiert einen Betrag von einer Fremdwährung nach EUR.

    Returns:
        tuple: (betrag_eur, original_string) oder (amount, None) wenn bereits EUR
    """
    if not currency or currency.upper() == 'EUR':
        return amount, None

    currency = currency.upper()
    rates = get_exchange_rates()

    if currency in rates:
        rate = rates[currency]
        amount_eur = round(amount * rate, 2)
        original_str = f"{amount:.2f} {currency}"
        return amount_eur, original_str
    else:
        print(f"⚠️  Unbekannte Währung: {currency} - keine Umrechnung")
        return amount, None


def process_currency_conversion(expense_data):
    """
    Verarbeitet einen Expense-Eintrag und konvertiert Fremdwährungen nach EUR.
    Modifiziert das Dict in-place und fügt Original-Währungsinfo zur Beschreibung hinzu.
    """
    betrag = expense_data.get('betrag', 0)
    waehrung = expense_data.get('waehrung', 'EUR')

    if not betrag or not waehrung:
        return expense_data

    betrag_eur, original_str = convert_to_eur(float(betrag), waehrung)

    if original_str:
        # Fremdwährung wurde konvertiert
        expense_data['betrag'] = betrag_eur
        expense_data['betrag_original'] = original_str
        expense_data['waehrung_original'] = waehrung
        expense_data['waehrung'] = 'EUR'

        # Original-Währung in Beschreibung einfügen
        beschreibung = expense_data.get('beschreibung', '')
        if beschreibung and original_str not in beschreibung:
            expense_data['beschreibung'] = f"{beschreibung} ({original_str})"

    return expense_data


# ============================================================================
# Export-Verzeichnis-Struktur
# ============================================================================

# Deutsche Monatsnamen
MONAT_NAMEN = {
    1: 'Januar', 2: 'Februar', 3: 'März', 4: 'April',
    5: 'Mai', 6: 'Juni', 7: 'Juli', 8: 'August',
    9: 'September', 10: 'Oktober', 11: 'November', 12: 'Dezember'
}

# Monat-Kurzformen für Parsing
MONAT_KURZ = {
    'jan': 1, 'feb': 2, 'mär': 3, 'mar': 3, 'apr': 4,
    'mai': 5, 'may': 5, 'jun': 6, 'jul': 7, 'aug': 8,
    'sep': 9, 'okt': 10, 'oct': 10, 'nov': 11, 'dez': 12, 'dec': 12
}

# Export-Basisverzeichnis
EXPORTS_DIR = os.environ.get('EXPORTS_DIR', os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exports'))


def parse_monat_string(monat_str):
    """
    Parst einen Monat-String (z.B. 'Nov 2025', 'November 2025', '11/2025')
    und gibt (Jahr, Monat) zurück.
    """
    if not monat_str:
        now = datetime.now()
        return now.year, now.month

    monat_str = monat_str.lower().strip()

    # Pattern 1: "Nov 2025" oder "November 2025"
    for kurz, num in MONAT_KURZ.items():
        if kurz in monat_str:
            # Jahr extrahieren
            import re
            year_match = re.search(r'(20\d{2})', monat_str)
            if year_match:
                return int(year_match.group(1)), num

    # Pattern 2: "11/2025" oder "11-2025" oder "11.2025"
    import re
    match = re.search(r'(\d{1,2})[/\-.]?(20\d{2})', monat_str)
    if match:
        return int(match.group(2)), int(match.group(1))

    # Fallback: aktueller Monat
    now = datetime.now()
    return now.year, now.month


def get_export_dir(monat_str):
    """
    Erstellt den Export-Pfad im Format: exports/2025/11_November/
    Gibt den Pfad zurück und erstellt das Verzeichnis falls nötig.
    """
    year, month = parse_monat_string(monat_str)
    month_name = MONAT_NAMEN.get(month, f'{month:02d}')

    export_path = os.path.join(EXPORTS_DIR, str(year), f'{month:02d}_{month_name}')
    os.makedirs(export_path, exist_ok=True)

    return export_path


# Archiv-Verzeichnis für verarbeitete Belege
ARCHIV_DIR = os.environ.get('ARCHIV_DIR', os.path.join(os.path.dirname(os.path.abspath(__file__)), 'belege', 'archiv'))


def get_archiv_dir(monat_str):
    """
    Erstellt den Archiv-Pfad im Format: belege/archiv/2025/11_November/
    Gibt den Pfad zurück und erstellt das Verzeichnis falls nötig.
    """
    year, month = parse_monat_string(monat_str)
    month_name = MONAT_NAMEN.get(month, f'{month:02d}')

    archiv_path = os.path.join(ARCHIV_DIR, str(year), f'{month:02d}_{month_name}')
    os.makedirs(archiv_path, exist_ok=True)

    return archiv_path


def archive_file(filepath, monat_str):
    """
    Verschiebt eine Datei ins Archiv-Verzeichnis.
    Gibt den neuen Pfad zurück oder None bei Fehler.
    """
    import shutil

    archiv_dir = get_archiv_dir(monat_str)
    filename = os.path.basename(filepath)
    target_path = os.path.join(archiv_dir, filename)

    # Bei Namenskonflikt: Nummer anhängen
    if os.path.exists(target_path):
        stem, suffix = os.path.splitext(filename)
        counter = 1
        while os.path.exists(target_path):
            target_path = os.path.join(archiv_dir, f"{stem}_{counter}{suffix}")
            counter += 1

    try:
        shutil.move(str(filepath), str(target_path))
        return target_path
    except Exception as e:
        print(f"⚠️  Archivierung fehlgeschlagen für {filename}: {e}")
        return None


CATEGORIES = {
    'fahrtkosten_kfz': 'Fahrtkosten mit priv. Kfz.',
    'fahrtkosten_pauschale': 'Fahrtkosten Öffentliche Verkehrsmittel',
    'bewirtung': 'Bewirtungskosten',
    'fachliteratur': 'Fachliteratur',
    'bueromaterial': 'Büromaterial',
    'telefonkosten': 'Telefonkosten',
    'software': 'Software',
    'getraenke': 'Getränke',
    'sonstiges': 'Sonstiges'
}


def get_anthropic_client():
    """Claude API Client erstellen"""
    api_key = os.getenv('ANTHROPIC_API_KEY')
    if not api_key:
        print("❌ ANTHROPIC_API_KEY nicht in .env gefunden!")
        return None
    return anthropic.Anthropic(api_key=api_key)


def process_receipt(filepath, client, cache=None, use_cache=True):
    """Einzelnen Beleg verarbeiten (mit Cache-Support)"""
    filename = os.path.basename(filepath).lower()
    images = []

    # Cache prüfen
    if use_cache and cache is not None:
        file_hash = get_file_hash(filepath)
        if file_hash in cache:
            cached_data = cache[file_hash]
            cached_data['datei'] = os.path.basename(filepath)
            cached_data['_cached'] = True
            return cached_data, None

    try:
        if filename.endswith('.pdf'):
            if not PDF_SUPPORT:
                return None, "PDF-Support nicht verfügbar"
            images = convert_from_path(filepath, dpi=300)
        elif filename.endswith(('.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.gif')):
            images = [Image.open(filepath)]
        else:
            return None, "Nicht unterstütztes Format"

        # OCR und Bild für AI vorbereiten
        full_text = ""
        first_image_base64 = None

        for idx, img in enumerate(images):
            # Erstes Bild für Claude Vision
            if idx == 0:
                img_for_ai = img.copy()
                img_for_ai.thumbnail((1568, 1568), Image.LANCZOS)
                if img_for_ai.mode in ('RGBA', 'P'):
                    img_for_ai = img_for_ai.convert('RGB')
                buffer = io.BytesIO()
                img_for_ai.save(buffer, format='JPEG', quality=85)
                first_image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')

            # OCR
            width, height = img.size
            if width < 2000:
                scale = 2000 / width
                img = img.resize((int(width * scale), int(height * scale)), Image.LANCZOS)
            img_gray = img.convert('L')
            text = pytesseract.image_to_string(img_gray, lang='deu+eng', config='--oem 3 --psm 3')
            full_text += text + "\n"

        full_text = full_text.strip()

        # Claude AI Analyse
        if not client:
            return None, "Kein API Client"

        prompt = """Analysiere diesen Beleg und extrahiere die Daten als JSON.
Antworte NUR mit dem JSON-Objekt.

Kategorien:
- fahrtkosten_kfz: Tankbelege, Benzin, Diesel
- fahrtkosten_pauschale: Fahrkarten, ÖPNV, Bahn, Bus
- bewirtung: Restaurant, Bar, Café
- fachliteratur: Bücher, Fachbücher
- bueromaterial: Bürobedarf
- telefonkosten: Telefon, Prepaid
- software: Software-Lizenzen
- getraenke: Getränke fürs Büro
- sonstiges: Parken, Taxi, Uber, Übernachtung, Hotel

WICHTIG für sonstiges - setze "typ" entsprechend:
- "Uber" wenn Uber, Bolt oder ähnliche Ride-Sharing-Dienste
- "Taxi" wenn klassisches Taxi
- "Parken" wenn Parkgebühren
- "Hotel" wenn Übernachtung
- "Sonstiges" für alles andere

JSON Format:
{
  "datum": "TT.MM.JJJJ",
  "betrag": 123.45,
  "waehrung": "EUR",
  "kategorie": "sonstiges",
  "typ": "Uber",
  "beschreibung": "Kurze Beschreibung",
  "anbieter": "Name des Geschäfts",
  "stadt": "Frankfurt",
  "distanz_km": 10.73
}

WICHTIG für Uber/Taxi:
- Extrahiere die Stadt aus der Anbieter-Adresse (z.B. "Frankfurt" aus "Albusstr. 17, 60313, Frankfurt")
- Extrahiere die Distanz in km wenn vorhanden (z.B. "Distanz: 10.73 km")
- Bei Uber Austria → stadt: "Wien"

WICHTIG zur Währung:
- Erkenne die Währung aus dem Beleg (EUR, CHF, USD, GBP, DKK, etc.)
- Verwende den GESAMTBETRAG inkl. MwSt/USt
- Bei Uber/Taxi: "Gesamtbetrag" ist der richtige Wert

Beleg-Text:
""" + full_text

        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1024,
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": "image/jpeg",
                            "data": first_image_base64
                        }
                    },
                    {"type": "text", "text": prompt}
                ]
            }]
        )

        response_text = message.content[0].text.strip()
        if response_text.startswith('```'):
            response_text = response_text.split('```')[1]
            if response_text.startswith('json'):
                response_text = response_text[4:]
        response_text = response_text.strip()

        data = json.loads(response_text)
        data['datei'] = os.path.basename(filepath)

        # File-Hash berechnen und mit speichern (für Beleg-Lookup)
        file_hash = get_file_hash(filepath)
        data['file_hash'] = file_hash

        # In Cache speichern (mit vollständigem Pfad für späteren Zugriff)
        if cache is not None:
            cache_data = {k: v for k, v in data.items() if k != 'datei'}
            cache_data['datei_pfad'] = str(filepath)  # Vollständiger Pfad
            cache[file_hash] = cache_data

        return data, None

    except Exception as e:
        return None, str(e)


def scan_folder(folder_path):
    """Ordner nach Belegen durchsuchen"""
    supported = ('.pdf', '.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.gif')
    files = []

    for f in Path(folder_path).iterdir():
        if f.is_file() and f.suffix.lower() in supported:
            files.append(str(f))

    return sorted(files)


def parse_datum(datum_str):
    """
    Parst ein Datum-String (z.B. '23.11.2025') und gibt ein datetime-Objekt zurück.
    Fallback: datetime.min für ungültige Daten (sortiert ans Ende).
    """
    if not datum_str:
        return datetime.min

    # Versuche verschiedene Formate
    for fmt in ['%d.%m.%Y', '%d.%m.%y', '%Y-%m-%d', '%d/%m/%Y']:
        try:
            return datetime.strptime(datum_str.strip(), fmt)
        except ValueError:
            continue

    return datetime.min


def sort_expenses_by_date(expenses):
    """Sortiert eine Liste von Expenses nach Datum (aufsteigend)."""
    return sorted(expenses, key=lambda e: parse_datum(e.get('datum', '')))


def export_excel(expenses, meta, output_path):
    """Excel-Export"""
    wb = Workbook()
    ws = wb.active
    ws.title = meta.get('monat', 'Spesen')

    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='333333')

    # Header
    ws['A1'] = f"Spesenabrechnung {meta.get('monat', '')}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = meta.get('name', '')
    ws['D1'] = f"Erstellt: {datetime.now().strftime('%d.%m.%Y')}"

    row = 4
    gesamt = 0

    # Nach Kategorien gruppieren
    by_category = {}
    for exp in expenses:
        cat = exp.get('kategorie', 'sonstiges')
        if cat not in by_category:
            by_category[cat] = []
        by_category[cat].append(exp)

    # Jede Kategorie nach Datum sortieren
    for cat in by_category:
        by_category[cat] = sort_expenses_by_date(by_category[cat])

    for cat_key, cat_name in CATEGORIES.items():
        cat_expenses = by_category.get(cat_key, [])
        if not cat_expenses:
            continue

        # Kategorie-Header
        ws.cell(row=row, column=1, value=cat_name)
        ws.cell(row=row, column=1).font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = header_fill

        headers = ['Datum', 'Beschreibung', 'Anbieter', 'Betrag']
        for i, h in enumerate(headers, 2):
            ws.cell(row=row, column=i, value=h)
            ws.cell(row=row, column=i).font = header_font
        row += 1

        cat_sum = 0
        for exp in cat_expenses:
            ws.cell(row=row, column=2, value=exp.get('datum', ''))
            ws.cell(row=row, column=3, value=exp.get('beschreibung', ''))
            ws.cell(row=row, column=4, value=exp.get('anbieter', ''))
            betrag = float(exp.get('betrag', 0) or 0)
            waehrung = exp.get('waehrung', 'EUR')
            ws.cell(row=row, column=5, value=f"{betrag:.2f} {waehrung}")
            cat_sum += betrag
            row += 1

        # Kategorie-Summe
        ws.cell(row=row, column=4, value="Summe:")
        ws.cell(row=row, column=4).font = Font(bold=True)
        ws.cell(row=row, column=5, value=f"{cat_sum:.2f} EUR")
        ws.cell(row=row, column=5).font = Font(bold=True)
        gesamt += cat_sum
        row += 2

    # Gesamtsumme
    ws.cell(row=row, column=4, value="GESAMT:")
    ws.cell(row=row, column=4).font = Font(size=12, bold=True)
    ws.cell(row=row, column=5, value=f"{gesamt:.2f} EUR")
    ws.cell(row=row, column=5).font = Font(size=12, bold=True)

    # Spaltenbreiten
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15

    wb.save(output_path)
    return gesamt


def get_db():
    """Datenbank-Verbindung herstellen"""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn


def save_to_database(expenses, meta):
    """Speichert die Abrechnung in der SQLite-Datenbank (fügt hinzu, überschreibt nicht)"""
    with get_db() as conn:
        # Prüfen ob Abrechnung für diesen Namen+Monat existiert
        existing = conn.execute(
            'SELECT id FROM abrechnungen WHERE name=? AND monat=?',
            (meta.get('name', ''), meta.get('monat', ''))
        ).fetchone()

        if existing:
            abrechnung_id = existing['id']
            # Update existierende Abrechnung (Datum aktualisieren)
            conn.execute('''
                UPDATE abrechnungen SET datum=?, updated_at=CURRENT_TIMESTAMP
                WHERE id=?
            ''', (meta.get('datum'), abrechnung_id))
            # NICHT mehr löschen - neue Belege werden hinzugefügt!
        else:
            # Neue Abrechnung erstellen
            cursor = conn.execute('''
                INSERT INTO abrechnungen (name, monat, datum)
                VALUES (?, ?, ?)
            ''', (meta.get('name', ''), meta.get('monat', ''), meta.get('datum')))
            abrechnung_id = cursor.lastrowid

        # Ausgaben nach Kategorien gruppiert speichern
        for exp in expenses:
            kategorie = exp.get('kategorie', 'sonstiges')
            if kategorie not in CATEGORIES:
                kategorie = 'sonstiges'

            # Daten für DB aufbereiten (Format wie Web-App erwartet)
            if kategorie == 'fahrtkosten_kfz':
                daten = {
                    'datum': exp.get('datum', ''),
                    'fahrstrecke': exp.get('beschreibung', ''),
                    'anlass': exp.get('anbieter', ''),
                    'km': 0  # Muss manuell eingetragen werden
                }
            elif kategorie == 'bewirtung':
                daten = {
                    'datum': exp.get('datum', ''),
                    'personen': exp.get('beschreibung', ''),
                    'betrag': exp.get('betrag', 0)
                }
            elif kategorie == 'sonstiges':
                # Typ aus AI-Antwort übernehmen, sonst Fallback
                typ = exp.get('typ', '')
                if not typ:
                    beschreibung = (exp.get('beschreibung', '') or '').lower()
                    anbieter = (exp.get('anbieter', '') or '').lower()

                    if 'uber' in beschreibung or 'uber' in anbieter or 'bolt' in anbieter:
                        typ = 'Uber'
                    elif 'taxi' in beschreibung or 'taxi' in anbieter:
                        typ = 'Taxi'
                    elif 'park' in beschreibung or 'park' in anbieter:
                        typ = 'Parken'
                    elif 'hotel' in beschreibung or 'hotel' in anbieter:
                        typ = 'Hotel'
                    elif 'verpflegung' in beschreibung or 'pauschale' in beschreibung:
                        typ = 'Verpflegungspauschale'
                    else:
                        typ = 'Sonstiges'

                # Ort zusammenbauen: Stadt + km wenn vorhanden
                betrag = exp.get('betrag', 0)
                waehrung = exp.get('waehrung', 'EUR')
                stadt = exp.get('stadt', '')
                distanz = exp.get('distanz_km')

                # Für Uber/Taxi: kompaktes Format "Stadt (X.XX km)"
                if typ in ('Uber', 'Taxi') and (stadt or distanz):
                    if stadt and distanz:
                        ort = f"{stadt} ({distanz} km)"
                    elif distanz:
                        ort = f"{distanz} km"
                    elif stadt:
                        ort = stadt
                    else:
                        ort = 'Fahrt'
                else:
                    ort = f"{exp.get('beschreibung', '')} - {exp.get('anbieter', '')}"

                # Fremdwährung anzeigen
                if waehrung and waehrung != 'EUR':
                    ort = f"{ort} ({betrag} {waehrung})"

                daten = {
                    'datum': exp.get('datum', ''),
                    'typ': typ,
                    'ort': ort,
                    'betrag': betrag
                }
            else:
                # Standard für andere Kategorien
                daten = {
                    'datum': exp.get('datum', ''),
                    'beschreibung': f"{exp.get('beschreibung', '')} ({exp.get('anbieter', '')})",
                    'betrag': exp.get('betrag', 0)
                }

            # File-Hash für Beleg-Lookup hinzufügen (falls vorhanden)
            if exp.get('file_hash'):
                daten['file_hash'] = exp.get('file_hash')

            conn.execute('''
                INSERT INTO ausgaben (abrechnung_id, kategorie, daten)
                VALUES (?, ?, ?)
            ''', (abrechnung_id, kategorie, json.dumps(daten)))

        conn.commit()
        return abrechnung_id


def export_pdf(expenses, meta, output_path):
    """PDF-Export"""
    doc = SimpleDocTemplate(output_path, pagesize=A4,
                           leftMargin=15*mm, rightMargin=15*mm,
                           topMargin=15*mm, bottomMargin=15*mm)

    styles = getSampleStyleSheet()
    elements = []

    # Titel
    title_style = ParagraphStyle('Title', parent=styles['Heading1'],
                                  fontSize=16, textColor=colors.HexColor('#333333'))
    elements.append(Paragraph(f"Spesenabrechnung {meta.get('monat', '')}", title_style))
    elements.append(Paragraph(f"Name: {meta.get('name', '')}", styles['Normal']))
    elements.append(Paragraph(f"Erstellt: {datetime.now().strftime('%d.%m.%Y')}", styles['Normal']))
    elements.append(Spacer(1, 10*mm))

    gesamt = 0

    # Nach Kategorien gruppieren
    by_category = {}
    for exp in expenses:
        cat = exp.get('kategorie', 'sonstiges')
        if cat not in by_category:
            by_category[cat] = []
        by_category[cat].append(exp)

    # Jede Kategorie nach Datum sortieren
    for cat in by_category:
        by_category[cat] = sort_expenses_by_date(by_category[cat])

    for cat_key, cat_name in CATEGORIES.items():
        cat_expenses = by_category.get(cat_key, [])
        if not cat_expenses:
            continue

        elements.append(Paragraph(cat_name, styles['Heading2']))

        table_data = [['Datum', 'Beschreibung', 'Anbieter', 'Betrag']]
        cat_sum = 0

        for exp in cat_expenses:
            betrag = float(exp.get('betrag', 0) or 0)
            waehrung = exp.get('waehrung', 'EUR')
            table_data.append([
                exp.get('datum', ''),
                exp.get('beschreibung', '')[:40],
                exp.get('anbieter', '')[:25],
                f"{betrag:.2f} {waehrung}"
            ])
            cat_sum += betrag

        table_data.append(['', '', 'Summe:', f"{cat_sum:.2f} EUR"])
        gesamt += cat_sum

        t = Table(table_data, colWidths=[25*mm, 70*mm, 50*mm, 30*mm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#333333')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (-1, 1), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 5*mm))

    # Gesamtsumme
    elements.append(Spacer(1, 10*mm))
    total_table = Table([['Gesamtsumme', f"{gesamt:.2f} EUR"]], colWidths=[145*mm, 30*mm])
    total_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#333333')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),
    ]))
    elements.append(total_table)

    doc.build(elements)
    return gesamt


def main():
    parser = argparse.ArgumentParser(
        description='Spesen CLI - Automatische Spesenabrechnung aus Belegen',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Beispiele:
  python cli.py /pfad/zu/belegen
  python cli.py /pfad/zu/belegen --name "Max Mustermann" --monat "Dez 2025"
  python cli.py /pfad/zu/belegen --output abrechnung.xlsx --format excel
  python cli.py belege/inbox --monat "Dez 2025" --archive

Inbox/Archiv Workflow:
  1. Belege in belege/inbox/ ablegen
  2. python cli.py belege/inbox --monat "Dez 2025" --archive
  3. Belege werden nach belege/archiv/2025/12_Dezember/ verschoben
        """
    )

    parser.add_argument('folder', help='Ordner mit Belegen (PDF, JPG, PNG)')
    parser.add_argument('--name', '-n', default='', help='Name für die Abrechnung')
    parser.add_argument('--monat', '-m', default='', help='Monat/Zeitraum (z.B. "Nov 2025")')
    parser.add_argument('--output', '-o', help='Ausgabedatei (Standard: spesen_DATUM.xlsx)')
    parser.add_argument('--format', '-f', choices=['excel', 'pdf', 'both', 'json'],
                        default='both', help='Ausgabeformat (Standard: both)')
    parser.add_argument('--no-db', action='store_true',
                        help='NICHT in Datenbank speichern (Standard: speichert in DB)')
    parser.add_argument('--no-cache', action='store_true',
                        help='Cache ignorieren und alle Belege neu verarbeiten')
    parser.add_argument('--archive', '-a', action='store_true',
                        help='Verarbeitete Belege ins Archiv verschieben (belege/archiv/Jahr/Monat/)')
    parser.add_argument('--verbose', '-v', action='store_true', help='Ausführliche Ausgabe')

    args = parser.parse_args()

    # Ordner prüfen
    if not os.path.isdir(args.folder):
        print(f"❌ Ordner nicht gefunden: {args.folder}")
        sys.exit(1)

    # Belege finden
    print(f"\n📁 Scanne Ordner: {args.folder}")
    files = scan_folder(args.folder)

    if not files:
        print("❌ Keine Belege gefunden (PDF, JPG, PNG)")
        sys.exit(1)

    print(f"📄 {len(files)} Belege gefunden")

    # Cache laden
    use_cache = not args.no_cache
    cache = load_cache() if use_cache else {}
    cached_count = 0

    if use_cache and cache:
        print(f"📦 Cache geladen ({len(cache)} Einträge)")
    print()

    # Claude Client
    client = get_anthropic_client()
    if not client:
        sys.exit(1)

    # Belege verarbeiten
    expenses = []
    errors = []
    processed_files = []  # Erfolgreich verarbeitete Dateien für Archivierung

    for i, filepath in enumerate(files, 1):
        filename = os.path.basename(filepath)
        print(f"[{i}/{len(files)}] Verarbeite: {filename}", end=" ", flush=True)

        data, error = process_receipt(filepath, client, cache=cache, use_cache=use_cache)

        if data:
            is_cached = data.pop('_cached', False)
            cache_indicator = " 📦" if is_cached else ""
            if is_cached:
                cached_count += 1

            # Original-Währung für Anzeige merken
            original_waehrung = data.get('waehrung', 'EUR')
            original_betrag = data.get('betrag', 0)

            # Fremdwährung nach EUR konvertieren
            if original_waehrung and original_waehrung.upper() != 'EUR':
                process_currency_conversion(data)
                conversion_indicator = f" → {data.get('betrag', '?')} EUR"
            else:
                conversion_indicator = ""

            expenses.append(data)
            processed_files.append(filepath)  # Für Archivierung merken
            print(f"✅ {original_betrag} {original_waehrung} - {data.get('kategorie', '?')}{conversion_indicator}{cache_indicator}")
            if args.verbose:
                print(f"         → {data.get('beschreibung', '')} ({data.get('anbieter', '')})")
        else:
            errors.append((filename, error))
            print(f"❌ {error}")

    # Cache speichern
    if use_cache:
        save_cache(cache)

    print(f"\n{'='*60}")
    print(f"✅ Erfolgreich: {len(expenses)} Belege", end="")
    if cached_count > 0:
        print(f" ({cached_count} aus Cache)")
    else:
        print()
    if errors:
        print(f"❌ Fehler: {len(errors)} Belege")

    if not expenses:
        print("\n❌ Keine Belege verarbeitet - Abbruch")
        sys.exit(1)

    # Meta-Daten
    meta = {
        'name': args.name,
        'monat': args.monat or datetime.now().strftime('%b %Y'),
        'datum': datetime.now().strftime('%d.%m.%Y')
    }

    # Gesamtsumme berechnen
    total = sum(float(e.get('betrag', 0) or 0) for e in expenses)
    print(f"\n💰 Gesamtsumme: {total:.2f} EUR")

    # In Datenbank speichern (Standard: ja, außer --no-db)
    if not args.no_db:
        try:
            abrechnung_id = save_to_database(expenses, meta)
            print(f"\n💾 In Datenbank gespeichert (ID: {abrechnung_id})")
            print(f"   → In Web-App sichtbar unter http://localhost:5000")
        except Exception as e:
            print(f"\n⚠️  Datenbank-Fehler: {e}")

    # Export - organisiert nach Jahr/Monat
    export_dir = get_export_dir(meta.get('monat', ''))
    monat_safe = meta.get('monat', 'Export').replace(' ', '_').replace('/', '-')

    # Dateinamen: Spesen_Nov_2025.xlsx
    if args.output:
        # Benutzer hat eigenen Pfad angegeben
        base_output = args.output
    else:
        # Automatischer Pfad: exports/2025/11_November/Spesen_Nov_2025
        base_output = os.path.join(export_dir, f"Spesen_{monat_safe}")

    if args.format in ('excel', 'both'):
        excel_path = base_output if base_output.endswith('.xlsx') else f"{base_output}.xlsx"
        export_excel(expenses, meta, excel_path)
        print(f"\n📊 Excel exportiert: {excel_path}")

    if args.format in ('pdf', 'both'):
        pdf_path = base_output if base_output.endswith('.pdf') else f"{base_output}.pdf"
        export_pdf(expenses, meta, pdf_path)
        print(f"📄 PDF exportiert: {pdf_path}")

    if args.format == 'json':
        json_path = base_output if base_output.endswith('.json') else f"{base_output}.json"
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump({'meta': meta, 'expenses': expenses, 'total': total}, f,
                     ensure_ascii=False, indent=2)
        print(f"\n📋 JSON exportiert: {json_path}")

    # Archivierung (optional)
    if args.archive and processed_files:
        print(f"\n📦 Archiviere {len(processed_files)} Belege...")
        archiv_dir = get_archiv_dir(meta.get('monat', ''))
        archived_count = 0
        for filepath in processed_files:
            result = archive_file(filepath, meta.get('monat', ''))
            if result:
                archived_count += 1
                if args.verbose:
                    print(f"   → {os.path.basename(filepath)}")
        print(f"✅ {archived_count} Belege archiviert nach: {archiv_dir}")

    print(f"\n✨ Fertig!")


if __name__ == '__main__':
    main()
