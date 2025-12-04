from flask import Flask, render_template, request, jsonify, send_file
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from cryptography.fernet import Fernet
from dotenv import load_dotenv
from PIL import Image
import anthropic
import pytesseract
import base64
import hashlib
import io
import json
import os
import re
import sqlite3

# Load environment variables
load_dotenv()

# Optional: pdf2image for PDF support
try:
    from pdf2image import convert_from_bytes
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

app = Flask(__name__)

# Data directory for storing files
DATA_DIR = os.environ.get('DATA_DIR', os.path.join(os.path.dirname(__file__), 'data'))
os.makedirs(DATA_DIR, exist_ok=True)

# Export directory (Jahr/Monat Struktur)
EXPORTS_DIR = os.environ.get('EXPORTS_DIR', os.path.join(os.path.dirname(__file__), 'exports'))

# Cache file for parsed receipts (same format as CLI)
CACHE_FILE = os.path.join(DATA_DIR, '.beleg_cache.json')

# Deutsche Monatsnamen für Export-Ordner
MONAT_NAMEN = {
    1: 'Januar', 2: 'Februar', 3: 'März', 4: 'April',
    5: 'Mai', 6: 'Juni', 7: 'Juli', 8: 'August',
    9: 'September', 10: 'Oktober', 11: 'November', 12: 'Dezember'
}

# Monat-Kurzformen für Parsing
MONAT_KURZ = {
    'jan': 1, 'feb': 2, 'mär': 3, 'mar': 3, 'apr': 4,
    'mai': 5, 'may': 5, 'jun': 6, 'jul': 7, 'aug': 8,
    'sep': 9, 'okt': 10, 'oct': 10, 'nov': 11, 'dez': 12, 'dec': 12
}


def parse_monat_string(monat_str):
    """Parst einen Monat-String und gibt (Jahr, Monat) zurück."""
    if not monat_str:
        now = datetime.now()
        return now.year, now.month

    monat_str = monat_str.lower().strip()

    # Pattern 1: "Nov 2025" oder "November 2025"
    for kurz, num in MONAT_KURZ.items():
        if kurz in monat_str:
            year_match = re.search(r'(20\d{2})', monat_str)
            if year_match:
                return int(year_match.group(1)), num

    # Pattern 2: "11/2025" oder "11-2025"
    match = re.search(r'(\d{1,2})[/\-.]?(20\d{2})', monat_str)
    if match:
        return int(match.group(2)), int(match.group(1))

    now = datetime.now()
    return now.year, now.month


def get_export_dir(monat_str, subfolder=None):
    """Erstellt Export-Pfad: exports/2025/11_November/ (optional mit Unterordner)"""
    year, month = parse_monat_string(monat_str)
    month_name = MONAT_NAMEN.get(month, f'{month:02d}')

    if subfolder:
        export_path = os.path.join(EXPORTS_DIR, str(year), f'{month:02d}_{month_name}', subfolder)
    else:
        export_path = os.path.join(EXPORTS_DIR, str(year), f'{month:02d}_{month_name}')

    os.makedirs(export_path, exist_ok=True)
    return export_path


def parse_datum(datum_str):
    """
    Parst ein Datum-String (z.B. '23.11.2025') und gibt ein datetime-Objekt zurück.
    Fallback: datetime.min für ungültige Daten (sortiert ans Ende).
    """
    if not datum_str:
        return datetime.min

    # Versuche verschiedene Formate
    for fmt in ['%d.%m.%Y', '%d.%m.%y', '%Y-%m-%d', '%d/%m/%Y']:
        try:
            return datetime.strptime(datum_str.strip(), fmt)
        except ValueError:
            continue

    return datetime.min


def sort_expenses_by_date(expenses):
    """Sortiert eine Liste von Expenses nach Datum (aufsteigend)."""
    return sorted(expenses, key=lambda e: parse_datum(e.get('datum', '')))


def get_content_hash(content):
    """Berechnet MD5-Hash des Dateiinhalts für Cache-Key"""
    hasher = hashlib.md5()
    hasher.update(content)
    return hasher.hexdigest()


def load_cache():
    """Lädt den Cache aus der JSON-Datei"""
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            return {}
    return {}


def save_cache(cache):
    """Speichert den Cache in die JSON-Datei"""
    try:
        with open(CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except IOError as e:
        print(f"Cache konnte nicht gespeichert werden: {e}")


# Encryption key - aus Umgebungsvariable oder Fallback auf Datei
def get_encryption_key():
    # Priorität 1: Umgebungsvariable
    env_key = os.environ.get('ENCRYPTION_KEY')
    if env_key:
        # Key kann als base64-String in .env gespeichert sein
        return env_key.encode() if isinstance(env_key, str) else env_key

    # Priorität 2: Bestehende Key-Datei (für Abwärtskompatibilität)
    key_file = os.path.join(DATA_DIR, 'secret.key')
    if os.path.exists(key_file):
        with open(key_file, 'rb') as f:
            return f.read()

    # Priorität 3: Neuen Key generieren und Warnung ausgeben
    key = Fernet.generate_key()
    print(f"⚠️  WARNUNG: Kein ENCRYPTION_KEY in .env gefunden!")
    print(f"   Generierter Key (bitte in .env speichern):")
    print(f"   ENCRYPTION_KEY={key.decode()}")

    # Key auch in Datei speichern als Backup
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(key_file, 'wb') as f:
        f.write(key)

    return key

ENCRYPTION_KEY = get_encryption_key()
cipher = Fernet(ENCRYPTION_KEY)

def encrypt_data(data):
    if not data:
        return None
    return cipher.encrypt(data.encode()).decode()

def decrypt_data(encrypted_data):
    if not encrypted_data:
        return None
    return cipher.decrypt(encrypted_data.encode()).decode()

DATABASE = os.path.join(DATA_DIR, 'spesen.db')

def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db() as conn:
        conn.execute('''
            CREATE TABLE IF NOT EXISTS abrechnungen (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                monat TEXT NOT NULL,
                datum TEXT,
                konto TEXT,
                blz TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(name, monat)
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS ausgaben (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                abrechnung_id INTEGER NOT NULL,
                kategorie TEXT NOT NULL,
                daten TEXT NOT NULL,
                FOREIGN KEY (abrechnung_id) REFERENCES abrechnungen(id) ON DELETE CASCADE
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS einstellungen (
                id INTEGER PRIMARY KEY CHECK (id = 1),
                name TEXT,
                iban_encrypted TEXT,
                bic_encrypted TEXT,
                bank TEXT,
                unterschrift_base64 TEXT,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS personen (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                firma TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.commit()

init_db()

# Verpflegungspauschalen 2025
VERPFLEGUNGSPAUSCHALEN = {
    'Deutschland': {'halbtag': 14.00, 'ganztag': 28.00},
    'Belgien': {'halbtag': 40.00, 'ganztag': 59.00},
    'Bulgarien': {'halbtag': 15.00, 'ganztag': 22.00},
    'Dänemark': {'halbtag': 50.00, 'ganztag': 75.00},
    'Finnland': {'halbtag': 36.00, 'ganztag': 54.00},
    'Frankreich': {'halbtag': 39.00, 'ganztag': 58.00},
    'Griechenland': {'halbtag': 27.00, 'ganztag': 40.00},
    'Großbritannien': {'halbtag': 38.00, 'ganztag': 57.00},
    'Irland': {'halbtag': 39.00, 'ganztag': 58.00},
    'Italien': {'halbtag': 28.00, 'ganztag': 42.00},
    'Kroatien': {'halbtag': 31.00, 'ganztag': 46.00},
    'Niederlande': {'halbtag': 32.00, 'ganztag': 47.00},
    'Österreich': {'halbtag': 33.00, 'ganztag': 50.00},
    'Polen': {'halbtag': 27.00, 'ganztag': 40.00},
    'Portugal': {'halbtag': 24.00, 'ganztag': 36.00},
    'Schweden': {'halbtag': 44.00, 'ganztag': 66.00},
    'Schweiz': {'halbtag': 43.00, 'ganztag': 64.00},
    'Spanien': {'halbtag': 28.00, 'ganztag': 42.00},
    'Tschechien': {'halbtag': 24.00, 'ganztag': 35.00},
    'USA': {'halbtag': 40.00, 'ganztag': 59.00},
}

CATEGORIES = {
    'fahrtkosten_kfz': {'name': 'Fahrtkosten mit priv. Kfz.', 'fields': ['datum', 'fahrstrecke', 'anlass', 'km'], 'rate': 0.30},
    'fahrtkosten_pauschale': {'name': 'Fahrtkosten Öffentliche Verkehrsmittel', 'fields': ['monat', 'beschreibung', 'betrag']},
    'bewirtung': {'name': 'Bewirtungskosten', 'fields': ['datum', 'personen', 'betrag']},
    'fachliteratur': {'name': 'Fachliteratur', 'fields': ['datum', 'beschreibung', 'betrag']},
    'bueromaterial': {'name': 'Büromaterial', 'fields': ['datum', 'beschreibung', 'betrag']},
    'telefonkosten': {'name': 'Telefonkosten', 'fields': ['datum', 'beschreibung', 'betrag']},
    'software': {'name': 'Software', 'fields': ['datum', 'beschreibung', 'betrag']},
    'getraenke': {'name': 'Getränke', 'fields': ['datum', 'beschreibung', 'betrag']},
    'sonstiges': {'name': 'Sonstiges', 'fields': ['datum', 'typ', 'ort', 'betrag'], 
                  'types': ['Parken', 'Taxi', 'Verpflegungspauschale', 'Uber', 'Sonstiges']}
}

@app.route('/')
def index():
    return render_template('index.html', categories=CATEGORIES, verpflegungspauschalen=VERPFLEGUNGSPAUSCHALEN)

@app.route('/health')
def health():
    """Health check endpoint for container orchestration."""
    try:
        # Test database connection
        with get_db() as conn:
            conn.execute('SELECT 1')
        return jsonify({
            'status': 'healthy',
            'database': 'connected',
            'timestamp': datetime.now().isoformat()
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 503

@app.route('/api/verpflegungspauschalen')
def get_verpflegungspauschalen():
    return jsonify(VERPFLEGUNGSPAUSCHALEN)

# Claude API Client
def get_anthropic_client():
    api_key = os.getenv('ANTHROPIC_API_KEY')
    print(f"[DEBUG] API Key vorhanden: {bool(api_key)}, Länge: {len(api_key) if api_key else 0}")
    if not api_key:
        return None
    return anthropic.Anthropic(api_key=api_key)

# Beleg-Parser mit Claude AI
def extract_receipt_data_with_ai(text, image_base64=None):
    """Extrahiert Daten aus Beleg mit Claude AI"""
    client = get_anthropic_client()

    if not client:
        # Fallback auf einfache Regex wenn kein API Key
        return extract_receipt_data_fallback(text)

    prompt = """Analysiere diesen Beleg/Quittung und extrahiere die folgenden Informationen.
Antworte NUR mit einem JSON-Objekt, ohne zusätzlichen Text.

Kategorien zur Auswahl:
- fahrtkosten_kfz: Tankbelege, Benzin, Diesel
- fahrtkosten_pauschale: Fahrkarten, Tickets (Bahn, Bus, ÖPNV)
- bewirtung: Restaurant, Bar, Café, Bewirtungskosten
- fachliteratur: Bücher, Fachbücher
- bueromaterial: Bürobedarf, Druckerpatronen
- telefonkosten: Telefon, Mobilfunk, Prepaid-Aufladungen
- software: Software-Lizenzen, Abos
- getraenke: Getränke fürs Büro
- sonstiges: Parken, Taxi, Übernachtung, alles andere

JSON Format:
{
  "datum": "TT.MM.JJJJ",
  "betrag": 123.45,
  "waehrung": "EUR",
  "kategorie": "bewirtung",
  "beschreibung": "Kurze Beschreibung",
  "anbieter": "Name des Geschäfts/Restaurants"
}

Wichtig:
- Datum im deutschen Format TT.MM.JJJJ
- Betrag als Zahl (nicht als String)
- Bei handschriftlichen Beträgen: bestmöglich interpretieren
- Bei Bewirtung: Restaurant-Name als Beschreibung
- Bei unleserlichen Werten: null verwenden

Beleg-Text:
"""

    try:
        # Wenn wir ein Bild haben, nutze Vision
        if image_base64:
            message = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=1024,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "image",
                                "source": {
                                    "type": "base64",
                                    "media_type": "image/jpeg",
                                    "data": image_base64
                                }
                            },
                            {
                                "type": "text",
                                "text": prompt + text
                            }
                        ]
                    }
                ]
            )
        else:
            message = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=1024,
                messages=[
                    {
                        "role": "user",
                        "content": prompt + text
                    }
                ]
            )

        # Parse JSON response
        response_text = message.content[0].text.strip()

        # Entferne mögliche Markdown-Codeblocks
        if response_text.startswith('```'):
            response_text = response_text.split('```')[1]
            if response_text.startswith('json'):
                response_text = response_text[4:]
        response_text = response_text.strip()

        data = json.loads(response_text)

        return {
            'datum': data.get('datum'),
            'betrag': data.get('betrag'),
            'waehrung': data.get('waehrung', 'EUR'),
            'kategorie_vorschlag': data.get('kategorie'),
            'beschreibung': data.get('beschreibung') or data.get('anbieter'),
            'anbieter': data.get('anbieter'),
            'raw_text': text
        }

    except Exception as e:
        print(f"Claude API Fehler: {e}")
        # Fallback auf Regex
        return extract_receipt_data_fallback(text)

def extract_receipt_data_fallback(text):
    """Einfacher Fallback wenn keine AI verfügbar"""
    result = {
        'datum': None,
        'betrag': None,
        'beschreibung': None,
        'kategorie_vorschlag': None,
        'raw_text': text
    }

    # Einfache Datum-Erkennung
    date_match = re.search(r'(\d{1,2})[./](\d{1,2})[./](\d{2,4})', text)
    if date_match:
        day, month, year = date_match.groups()
        if len(year) == 2:
            year = '20' + year
        result['datum'] = f"{int(day):02d}.{int(month):02d}.{year}"

    # Einfache Betrag-Erkennung
    amount_match = re.search(r'(\d+)[,.](\d{2})\s*(?:EUR|€)', text, re.IGNORECASE)
    if amount_match:
        result['betrag'] = float(f"{amount_match.group(1)}.{amount_match.group(2)}")

    return result

@app.route('/api/parse-beleg', methods=['POST'])
def parse_beleg():
    """Parst einen hochgeladenen Beleg (Bild oder PDF) mit Claude AI"""
    if 'beleg' not in request.files:
        return jsonify({'error': 'Keine Datei hochgeladen'}), 400

    file = request.files['beleg']
    if file.filename == '':
        return jsonify({'error': 'Keine Datei ausgewählt'}), 400

    try:
        filename = file.filename.lower()
        original_filename = file.filename

        # Dateiinhalt lesen für Hash-Berechnung
        file_content = file.read()
        file.seek(0)  # Zurücksetzen für weitere Verarbeitung

        # Cache prüfen
        content_hash = get_content_hash(file_content)
        cache = load_cache()

        if content_hash in cache:
            # Aus Cache laden
            cached_data = cache[content_hash].copy()
            cached_data['_cached'] = True
            return jsonify({
                'success': True,
                'data': cached_data
            })

        images = []
        first_image_base64 = None

        if filename.endswith('.pdf'):
            if not PDF_SUPPORT:
                return jsonify({'error': 'PDF-Support nicht verfügbar. Bitte poppler installieren.'}), 400
            # PDF zu Bildern konvertieren
            images = convert_from_bytes(file_content, dpi=300)
        elif filename.endswith(('.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.gif')):
            # Bild direkt laden
            images = [Image.open(io.BytesIO(file_content))]
        else:
            return jsonify({'error': 'Nicht unterstütztes Dateiformat. Erlaubt: PDF, PNG, JPG, TIFF'}), 400

        # OCR auf allen Seiten/Bildern durchführen
        full_text = ""
        for idx, img in enumerate(images):
            # Original für AI-Vision speichern (erstes Bild)
            if idx == 0:
                # Bild für Claude vorbereiten (max 1568px, als JPEG)
                img_for_ai = img.copy()
                img_for_ai.thumbnail((1568, 1568), Image.LANCZOS)
                if img_for_ai.mode in ('RGBA', 'P'):
                    img_for_ai = img_for_ai.convert('RGB')
                buffer = io.BytesIO()
                img_for_ai.save(buffer, format='JPEG', quality=85)
                first_image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')

            # Bild für OCR optimieren
            width, height = img.size
            if width < 2000:
                scale = 2000 / width
                img = img.resize((int(width * scale), int(height * scale)), Image.LANCZOS)

            # Graustufen
            img_gray = img.convert('L')

            # OCR
            custom_config = r'--oem 3 --psm 3'
            text = pytesseract.image_to_string(img_gray, lang='deu+eng', config=custom_config)
            full_text += text + "\n"

        full_text = full_text.strip()

        # Daten mit Claude AI extrahieren (mit Bild für bessere Erkennung)
        extracted = extract_receipt_data_with_ai(full_text, first_image_base64)

        # In Cache speichern
        cache[content_hash] = extracted
        save_cache(cache)

        return jsonify({
            'success': True,
            'data': extracted
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Fehler beim Parsen: {str(e)}'}), 500

# API: Einstellungen laden
@app.route('/api/einstellungen', methods=['GET'])
def get_einstellungen():
    with get_db() as conn:
        row = conn.execute('SELECT * FROM einstellungen WHERE id = 1').fetchone()
        if row:
            return jsonify({
                'name': row['name'],
                'iban': decrypt_data(row['iban_encrypted']),
                'bic': decrypt_data(row['bic_encrypted']),
                'bank': row['bank']
            })
        return jsonify({'name': '', 'iban': '', 'bic': '', 'bank': ''})

# API: Einstellungen speichern
@app.route('/api/einstellungen', methods=['POST'])
def save_einstellungen():
    data = request.json

    iban_encrypted = encrypt_data(data.get('iban', ''))
    bic_encrypted = encrypt_data(data.get('bic', ''))

    with get_db() as conn:
        conn.execute('''
            INSERT INTO einstellungen (id, name, iban_encrypted, bic_encrypted, bank, unterschrift_base64)
            VALUES (1, ?, ?, ?, ?, ?)
            ON CONFLICT(id) DO UPDATE SET
                name = excluded.name,
                iban_encrypted = excluded.iban_encrypted,
                bic_encrypted = excluded.bic_encrypted,
                bank = excluded.bank,
                unterschrift_base64 = COALESCE(excluded.unterschrift_base64, unterschrift_base64),
                updated_at = CURRENT_TIMESTAMP
        ''', (data.get('name', ''), iban_encrypted, bic_encrypted, data.get('bank', ''), data.get('unterschrift_base64')))
        conn.commit()

    return jsonify({'success': True})

@app.route('/api/unterschrift', methods=['POST'])
def save_unterschrift():
    """Speichert die Unterschrift als Base64-PNG."""
    data = request.json
    unterschrift = data.get('unterschrift_base64', '')

    with get_db() as conn:
        conn.execute('''
            INSERT INTO einstellungen (id, unterschrift_base64)
            VALUES (1, ?)
            ON CONFLICT(id) DO UPDATE SET
                unterschrift_base64 = excluded.unterschrift_base64,
                updated_at = CURRENT_TIMESTAMP
        ''', (unterschrift,))
        conn.commit()

    return jsonify({'success': True})

@app.route('/api/unterschrift', methods=['GET'])
def get_unterschrift():
    """Gibt die Unterschrift als Base64 zurück."""
    with get_db() as conn:
        row = conn.execute('SELECT unterschrift_base64 FROM einstellungen WHERE id = 1').fetchone()
        if row and row['unterschrift_base64']:
            return jsonify({'unterschrift_base64': row['unterschrift_base64']})
        return jsonify({'unterschrift_base64': None})

# API: Personen verwalten
@app.route('/api/personen', methods=['GET'])
def list_personen():
    """Liste aller gespeicherten Personen."""
    with get_db() as conn:
        rows = conn.execute('SELECT id, name, firma FROM personen ORDER BY name').fetchall()
        return jsonify([dict(row) for row in rows])

@app.route('/api/personen', methods=['POST'])
def add_person():
    """Fügt eine neue Person hinzu."""
    data = request.json
    name = data.get('name', '').strip()
    firma = data.get('firma', '').strip()

    if not name:
        return jsonify({'error': 'Name ist erforderlich'}), 400

    with get_db() as conn:
        cursor = conn.execute('INSERT INTO personen (name, firma) VALUES (?, ?)', (name, firma))
        conn.commit()
        return jsonify({'success': True, 'id': cursor.lastrowid})

@app.route('/api/personen/<int:person_id>', methods=['PUT'])
def update_person(person_id):
    """Aktualisiert eine Person."""
    data = request.json
    name = data.get('name', '').strip()
    firma = data.get('firma', '').strip()

    if not name:
        return jsonify({'error': 'Name ist erforderlich'}), 400

    with get_db() as conn:
        conn.execute('UPDATE personen SET name = ?, firma = ? WHERE id = ?', (name, firma, person_id))
        conn.commit()
        return jsonify({'success': True})

@app.route('/api/personen/<int:person_id>', methods=['DELETE'])
def delete_person(person_id):
    """Löscht eine Person."""
    with get_db() as conn:
        conn.execute('DELETE FROM personen WHERE id = ?', (person_id,))
        conn.commit()
        return jsonify({'success': True})

@app.route('/api/personen/import-vcf', methods=['POST'])
def import_vcf():
    """Importiert Personen aus VCF-Datei."""
    if 'vcf' not in request.files:
        return jsonify({'error': 'Keine VCF-Datei hochgeladen'}), 400

    vcf_file = request.files['vcf']
    if vcf_file.filename == '':
        return jsonify({'error': 'Keine Datei ausgewählt'}), 400

    try:
        content = vcf_file.read().decode('utf-8', errors='replace')

        # VCF Parser
        imported = []
        skipped = 0

        # VCF kann mehrere vCards enthalten
        vcards = content.split('BEGIN:VCARD')

        for vcard in vcards:
            if not vcard.strip():
                continue

            name = None
            firma = None

            # Parse vCard Felder
            for line in vcard.split('\n'):
                line = line.strip()

                # Name aus FN (Formatted Name) oder N
                if line.startswith('FN:') or line.startswith('FN;'):
                    # FN kann encoding haben: FN;CHARSET=UTF-8:Name
                    name = line.split(':', 1)[-1].strip()
                elif line.startswith('N:') or line.startswith('N;'):
                    # N Format: Nachname;Vorname;...
                    n_parts = line.split(':', 1)[-1].split(';')
                    if len(n_parts) >= 2:
                        nachname = n_parts[0].strip()
                        vorname = n_parts[1].strip()
                        if not name and (nachname or vorname):
                            name = f"{vorname} {nachname}".strip()

                # Organisation
                if line.startswith('ORG:') or line.startswith('ORG;'):
                    firma = line.split(':', 1)[-1].split(';')[0].strip()

                # Titel als Alternative für Firma
                if not firma and (line.startswith('TITLE:') or line.startswith('TITLE;')):
                    firma = line.split(':', 1)[-1].strip()

            if name:
                imported.append({'name': name, 'firma': firma or ''})
            else:
                skipped += 1

        # In Datenbank speichern (Duplikate vermeiden)
        added = 0
        with get_db() as conn:
            existing = {row['name'].lower() for row in conn.execute('SELECT name FROM personen').fetchall()}

            for person in imported:
                if person['name'].lower() not in existing:
                    conn.execute('INSERT INTO personen (name, firma) VALUES (?, ?)',
                               (person['name'], person['firma']))
                    existing.add(person['name'].lower())
                    added += 1
                else:
                    skipped += 1

            conn.commit()

        return jsonify({
            'success': True,
            'imported': added,
            'skipped': skipped,
            'total': len(imported)
        })

    except Exception as e:
        return jsonify({'error': f'Fehler beim Parsen: {str(e)}'}), 400

# API: Beleg per Hash abrufen
@app.route('/api/beleg/<file_hash>', methods=['GET'])
def get_beleg(file_hash):
    """Gibt den Beleg (PDF/Bild) per file_hash zurück."""
    import os
    from flask import send_file

    # Cache laden
    cache = {}
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE, 'r') as f:
            cache = json.load(f)

    if file_hash not in cache:
        return jsonify({'error': 'Beleg nicht im Cache gefunden'}), 404

    cache_entry = cache[file_hash]
    datei_pfad = cache_entry.get('datei_pfad')

    if not datei_pfad or not os.path.exists(datei_pfad):
        # Fallback: Dateiname in bekannten Ordnern suchen
        datei_name = cache_entry.get('datei')
        if datei_name:
            # Bekannte Beleg-Ordner (aus docker-compose)
            search_dirs = [
                '/data/belege',
                '/data/uber',
                os.path.expanduser('~/Documents/Scans'),
                os.path.expanduser('~/Desktop/Belege'),
                os.path.join(os.path.dirname(__file__), 'belege', 'archiv')
            ]
            for search_dir in search_dirs:
                if os.path.exists(search_dir):
                    for root, dirs, files in os.walk(search_dir):
                        if datei_name in files:
                            datei_pfad = os.path.join(root, datei_name)
                            break
                    if datei_pfad and os.path.exists(datei_pfad):
                        break

    if not datei_pfad or not os.path.exists(datei_pfad):
        return jsonify({'error': 'Beleg-Datei nicht gefunden', 'datei': cache_entry.get('datei')}), 404

    # MIME-Type bestimmen
    ext = os.path.splitext(datei_pfad)[1].lower()
    mime_types = {
        '.pdf': 'application/pdf',
        '.png': 'image/png',
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.tiff': 'image/tiff',
        '.bmp': 'image/bmp',
        '.gif': 'image/gif'
    }
    mime_type = mime_types.get(ext, 'application/octet-stream')

    return send_file(datei_pfad, mimetype=mime_type)

# API: Beleg-Info per Hash abrufen (ohne Datei)
@app.route('/api/beleg/<file_hash>/info', methods=['GET'])
def get_beleg_info(file_hash):
    """Gibt die Cache-Informationen zum Beleg zurück."""
    # Cache laden
    cache = {}
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE, 'r') as f:
            cache = json.load(f)

    if file_hash not in cache:
        return jsonify({'error': 'Beleg nicht im Cache gefunden'}), 404

    return jsonify(cache[file_hash])

# API: Liste aller Abrechnungen
@app.route('/api/abrechnungen', methods=['GET'])
def list_abrechnungen():
    with get_db() as conn:
        rows = conn.execute(
            'SELECT id, name, monat, datum, konto, blz, created_at FROM abrechnungen ORDER BY created_at DESC'
        ).fetchall()
        return jsonify([dict(row) for row in rows])

# API: Abrechnung laden
@app.route('/api/abrechnungen/<int:abrechnung_id>', methods=['GET'])
def get_abrechnung(abrechnung_id):
    with get_db() as conn:
        abr = conn.execute(
            'SELECT * FROM abrechnungen WHERE id = ?', (abrechnung_id,)
        ).fetchone()
        if not abr:
            return jsonify({'error': 'Nicht gefunden'}), 404

        ausgaben_rows = conn.execute(
            'SELECT kategorie, daten FROM ausgaben WHERE abrechnung_id = ?', (abrechnung_id,)
        ).fetchall()

        expenses = {cat: [] for cat in CATEGORIES.keys()}
        for row in ausgaben_rows:
            if row['kategorie'] in expenses:
                expenses[row['kategorie']].append(json.loads(row['daten']))

        # Jede Kategorie nach Datum sortieren
        for cat in expenses:
            expenses[cat] = sort_expenses_by_date(expenses[cat])

        return jsonify({
            'meta': {
                'id': abr['id'],
                'name': abr['name'],
                'monat': abr['monat'],
                'datum': abr['datum'],
                'konto': abr['konto'],
                'blz': abr['blz']
            },
            'expenses': expenses
        })

# API: Abrechnung speichern (neu oder update)
@app.route('/api/abrechnungen', methods=['POST'])
def save_abrechnung():
    data = request.json
    meta = data.get('meta', {})
    expenses = data.get('expenses', {})

    with get_db() as conn:
        # Prüfen ob Abrechnung existiert (per ID oder Name+Monat)
        abrechnung_id = meta.get('id')

        if abrechnung_id:
            # Update existierende
            conn.execute('''
                UPDATE abrechnungen SET name=?, monat=?, datum=?, konto=?, blz=?, updated_at=CURRENT_TIMESTAMP
                WHERE id=?
            ''', (meta.get('name'), meta.get('monat'), meta.get('datum'),
                  meta.get('konto'), meta.get('blz'), abrechnung_id))
        else:
            # Versuche existierende per Name+Monat zu finden
            existing = conn.execute(
                'SELECT id FROM abrechnungen WHERE name=? AND monat=?',
                (meta.get('name'), meta.get('monat'))
            ).fetchone()

            if existing:
                abrechnung_id = existing['id']
                conn.execute('''
                    UPDATE abrechnungen SET datum=?, konto=?, blz=?, updated_at=CURRENT_TIMESTAMP
                    WHERE id=?
                ''', (meta.get('datum'), meta.get('konto'), meta.get('blz'), abrechnung_id))
            else:
                # Neue Abrechnung erstellen
                cursor = conn.execute('''
                    INSERT INTO abrechnungen (name, monat, datum, konto, blz)
                    VALUES (?, ?, ?, ?, ?)
                ''', (meta.get('name'), meta.get('monat'), meta.get('datum'),
                      meta.get('konto'), meta.get('blz')))
                abrechnung_id = cursor.lastrowid

        # Alte Ausgaben löschen und neue einfügen
        conn.execute('DELETE FROM ausgaben WHERE abrechnung_id = ?', (abrechnung_id,))

        for kategorie, items in expenses.items():
            for item in items:
                conn.execute('''
                    INSERT INTO ausgaben (abrechnung_id, kategorie, daten)
                    VALUES (?, ?, ?)
                ''', (abrechnung_id, kategorie, json.dumps(item)))

        conn.commit()
        return jsonify({'success': True, 'id': abrechnung_id})

# API: Abrechnung löschen
@app.route('/api/abrechnungen/<int:abrechnung_id>', methods=['DELETE'])
def delete_abrechnung(abrechnung_id):
    with get_db() as conn:
        conn.execute('DELETE FROM ausgaben WHERE abrechnung_id = ?', (abrechnung_id,))
        conn.execute('DELETE FROM abrechnungen WHERE id = ?', (abrechnung_id,))
        conn.commit()
        return jsonify({'success': True})

@app.route('/export/excel', methods=['POST'])
def export_excel():
    data = request.json
    meta = data.get('meta', {})
    expenses = data.get('expenses', {})
    
    wb = Workbook()
    ws = wb.active
    ws.title = meta.get('monat', 'Spesen')
    
    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='333333')
    title_font = Font(color='333333', size=14, bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header
    ws['A1'] = meta.get('monat', '')
    ws['A1'].font = title_font
    ws['C1'] = meta.get('name', 'Olivier Dobberkau')
    ws['C1'].font = title_font
    ws['F1'] = f"Datum {meta.get('datum', datetime.now().strftime('%d.%m.%y'))}"
    ws['F1'].font = title_font
    ws['H1'] = 'Summen'
    ws['H1'].font = Font(size=14, bold=True)
    
    row = 3
    gesamt = 0
    beleg_nr = 1  # Fortlaufende Belegnummer

    for cat_key, cat_info in CATEGORIES.items():
        cat_expenses = expenses.get(cat_key, [])
        # Nach Datum sortieren
        cat_expenses = sort_expenses_by_date(cat_expenses)
        cat_sum = 0

        # Category header
        ws.cell(row=row, column=1, value=f"{list(CATEGORIES.keys()).index(cat_key) + 1}. {cat_info['name']}")
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=1).font = header_font

        if cat_key == 'fahrtkosten_kfz':
            headers = ['Nr.', 'Datum', 'Fahrstrecke', 'Anlaß', 'km', '0,3', 'Betrag €']
            for i, h in enumerate(headers, 2):
                ws.cell(row=row, column=i, value=h)
                ws.cell(row=row, column=i).fill = header_fill
                ws.cell(row=row, column=i).font = header_font

        row += 1
        start_row = row

        for exp in cat_expenses:
            # Belegnummer in erste Spalte (außer bei Sonstiges, da steht der Typ)
            if cat_key != 'sonstiges':
                ws.cell(row=row, column=2, value=beleg_nr)

            if cat_key == 'fahrtkosten_kfz':
                ws.cell(row=row, column=3, value=exp.get('datum', ''))
                ws.cell(row=row, column=4, value=exp.get('fahrstrecke', ''))
                ws.cell(row=row, column=5, value=exp.get('anlass', ''))
                km = float(exp.get('km', 0) or 0)
                ws.cell(row=row, column=6, value=km)
                betrag = km * 0.30
                ws.cell(row=row, column=8, value=betrag)
                cat_sum += betrag
            elif cat_key == 'sonstiges':
                ws.cell(row=row, column=1, value=f"{beleg_nr}. {exp.get('typ', '')}")
                ws.cell(row=row, column=2, value=exp.get('datum', ''))
                ws.cell(row=row, column=3, value=exp.get('ort', ''))
                betrag = float(exp.get('betrag', 0) or 0)
                ws.cell(row=row, column=7, value=betrag)
                cat_sum += betrag
            elif cat_key == 'bewirtung':
                ws.cell(row=row, column=3, value=exp.get('datum', ''))
                ws.cell(row=row, column=5, value=exp.get('personen', ''))
                betrag = float(exp.get('betrag', 0) or 0)
                ws.cell(row=row, column=8, value=betrag)
                cat_sum += betrag
            else:
                ws.cell(row=row, column=3, value=exp.get('datum', '') or exp.get('monat', ''))
                ws.cell(row=row, column=4, value=exp.get('beschreibung', ''))
                betrag = float(exp.get('betrag', 0) or 0)
                ws.cell(row=row, column=8, value=betrag)
                cat_sum += betrag
            beleg_nr += 1
            row += 1
        
        if not cat_expenses:
            row += 1
        
        # Sum for category
        ws.cell(row=row-1 if cat_expenses else row, column=8, value=cat_sum)
        ws.cell(row=row-1 if cat_expenses else row, column=8).number_format = '#,##0.00 €'
        gesamt += cat_sum
        row += 1
    
    # Total
    row += 1
    ws.cell(row=row, column=7, value='GESAMT')
    ws.cell(row=row, column=7).font = Font(bold=True)
    ws.cell(row=row, column=8, value=gesamt)
    ws.cell(row=row, column=8).font = Font(bold=True)
    ws.cell(row=row, column=8).number_format = '#,##0.00 €'
    
    # Bank details
    row += 2
    ws.cell(row=row, column=2, value=f"IBAN: {meta.get('iban', '')}")
    ws.cell(row=row, column=2).font = Font(color='333333')
    if meta.get('bic'):
        ws.cell(row=row, column=5, value=f"BIC: {meta.get('bic', '')}")
        ws.cell(row=row, column=5).font = Font(color='333333')
    ws.cell(row=row, column=7, value=meta.get('name', ''))
    ws.cell(row=row, column=7).font = Font(color='333333')
    
    # Column widths
    widths = [25, 12, 20, 30, 8, 8, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Spesen_{meta.get('monat', 'Export').replace(' ', '_')}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)

def generate_pdf_buffer(meta, expenses):
    """Generiert einen PDF-Buffer für die Spesenabrechnung."""
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, 
                           topMargin=15*mm, bottomMargin=15*mm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16,
                                  textColor=colors.HexColor('#333333'), spaceAfter=20)
    
    elements = []
    
    # Title
    title = f"Spesenabrechnung {meta.get('monat', '')} - {meta.get('name', 'Olivier Dobberkau')}"
    elements.append(Paragraph(title, title_style))
    elements.append(Paragraph(f"Datum: {meta.get('datum', datetime.now().strftime('%d.%m.%Y'))}", styles['Normal']))
    elements.append(Spacer(1, 10*mm))
    
    gesamt = 0
    beleg_nr = 1  # Fortlaufende Belegnummer über alle Kategorien

    for cat_key, cat_info in CATEGORIES.items():
        cat_expenses = expenses.get(cat_key, [])
        if not cat_expenses:
            continue

        # Nach Datum sortieren
        cat_expenses = sort_expenses_by_date(cat_expenses)

        cat_sum = 0
        elements.append(Paragraph(cat_info['name'], styles['Heading2']))

        # Build table data with Nr. column
        if cat_key == 'fahrtkosten_kfz':
            table_data = [['Nr.', 'Datum', 'Fahrstrecke', 'Anlaß', 'km', 'Betrag']]
            for exp in cat_expenses:
                km = float(exp.get('km', 0) or 0)
                betrag = km * 0.30
                cat_sum += betrag
                table_data.append([str(beleg_nr), exp.get('datum', ''), exp.get('fahrstrecke', ''),
                                  exp.get('anlass', ''), f"{km:.0f}", f"{betrag:.2f} €"])
                beleg_nr += 1
        elif cat_key == 'sonstiges':
            table_data = [['Nr.', 'Typ', 'Datum', 'Ort / Beschreibung', 'Betrag']]
            for exp in cat_expenses:
                betrag = float(exp.get('betrag', 0) or 0)
                cat_sum += betrag

                # Typ (Abkürzung für Verpflegungspauschale)
                typ = exp.get('typ', '')
                if typ == 'Verpflegungspauschale':
                    typ = 'VP'

                # Ort direkt übernehmen (wie in Web-App)
                ort = exp.get('ort', '')
                # Nur bei sehr langen Texten kürzen
                if len(ort) > 55:
                    ort = ort[:52] + '...'

                table_data.append([str(beleg_nr), typ, exp.get('datum', ''), ort, f"{betrag:.2f} €"])
                beleg_nr += 1
        elif cat_key == 'bewirtung':
            table_data = [['Nr.', 'Datum', 'Restaurant', 'Betrag']]
            for exp in cat_expenses:
                betrag = float(exp.get('betrag', 0) or 0)
                cat_sum += betrag
                # Restaurant-Name extrahieren (erster Teil vor " - ")
                personen = exp.get('personen', '')
                restaurant = personen.split(' - ')[0] if ' - ' in personen else personen
                # Auf max 45 Zeichen begrenzen
                if len(restaurant) > 45:
                    restaurant = restaurant[:42] + '...'
                table_data.append([str(beleg_nr), exp.get('datum', ''), restaurant, f"{betrag:.2f} €"])
                beleg_nr += 1
        else:
            table_data = [['Nr.', 'Datum', 'Beschreibung', 'Betrag']]
            for exp in cat_expenses:
                betrag = float(exp.get('betrag', 0) or 0)
                cat_sum += betrag
                table_data.append([str(beleg_nr), exp.get('datum', '') or exp.get('monat', ''),
                                  exp.get('beschreibung', ''), f"{betrag:.2f} €"])
                beleg_nr += 1

        table_data.append(['', '', '', 'Summe:', f"{cat_sum:.2f} €"] if len(table_data[0]) == 5
                         else ['', '', 'Summe:', f"{cat_sum:.2f} €"] if len(table_data[0]) == 4
                         else ['', '', '', '', 'Summe:', f"{cat_sum:.2f} €"])

        # Volle Seitenbreite: A4 = 210mm, minus 30mm Ränder = 180mm
        page_width = 180*mm
        num_cols = len(table_data[0])
        if cat_key == 'fahrtkosten_kfz':
            # Nr., Datum, Fahrstrecke, Anlaß, km, Betrag
            col_widths = [12*mm, 22*mm, 45*mm, 68*mm, 13*mm, 20*mm]
        elif cat_key == 'sonstiges':
            # Nr., Typ, Datum, Ort, Betrag
            col_widths = [12*mm, 25*mm, 22*mm, 96*mm, 25*mm]
        elif cat_key == 'bewirtung':
            # Nr., Datum, Restaurant, Betrag
            col_widths = [12*mm, 22*mm, 121*mm, 25*mm]
        else:
            # Nr., Datum, Beschreibung, Betrag
            col_widths = [12*mm, 22*mm, 121*mm, 25*mm]
        t = Table(table_data, colWidths=col_widths)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#333333')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (-1, 1), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 5*mm))
        gesamt += cat_sum
    
    # Total
    elements.append(Spacer(1, 10*mm))
    total_data = [['Gesamtsumme', f"{gesamt:.2f} €"]]
    total_table = Table(total_data, colWidths=[155*mm, 25*mm])
    total_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#333333')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 14),
        ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),
    ]))
    elements.append(total_table)
    
    # Bank details
    if meta.get('iban'):
        elements.append(Spacer(1, 15*mm))
        bank_parts = [f"IBAN: {meta.get('iban', '')}"]
        if meta.get('bic'):
            bank_parts.append(f"BIC: {meta.get('bic', '')}")
        bank_parts.append(meta.get('name', ''))
        bank_info = " | ".join(bank_parts)
        elements.append(Paragraph(bank_info, styles['Normal']))

    # Abkürzungsverzeichnis
    elements.append(Spacer(1, 10*mm))
    abbr_style = ParagraphStyle('Abbr', parent=styles['Normal'], fontSize=8, textColor=colors.grey)
    elements.append(Paragraph("<b>Abkürzungen:</b> VP = Verpflegungspauschale", abbr_style))

    doc.build(elements)
    output.seek(0)
    return output


@app.route('/export/pdf', methods=['POST'])
def export_pdf():
    data = request.json
    meta = data.get('meta', {})
    expenses = data.get('expenses', {})

    output = generate_pdf_buffer(meta, expenses)
    filename = f"Spesen_{meta.get('monat', 'Export').replace(' ', '_')}.pdf"
    return send_file(output, mimetype='application/pdf', as_attachment=True, download_name=filename)


@app.route('/export/bewirtungsbeleg', methods=['POST'])
def export_bewirtungsbeleg():
    """Generiert einen offiziellen Bewirtungsbeleg nach §4 Abs. 5 Nr. 2 EStG."""
    data = request.json

    # Daten aus dem Request
    datum = data.get('datum', '')
    restaurant = data.get('restaurant', '')
    ort = data.get('ort', '')
    betrag = float(data.get('betrag', 0) or 0)
    anlass = data.get('anlass', 'Geschäftliche Besprechung')
    bewirtende_person = data.get('bewirtende_person', '')
    teilnehmer = data.get('teilnehmer', [])  # Liste von {name, firma}
    unterschrift_base64 = data.get('unterschrift_base64', None)
    monat = data.get('monat', 'Unbekannt')
    beleg_nr = data.get('beleg_nr', None)  # Fortlaufende Belegnummer

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, leftMargin=20*mm, rightMargin=20*mm,
                           topMargin=20*mm, bottomMargin=20*mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18,
                                  textColor=colors.HexColor('#333333'), spaceAfter=15, alignment=1)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10,
                                     textColor=colors.grey, spaceAfter=20, alignment=1)
    label_style = ParagraphStyle('Label', parent=styles['Normal'], fontSize=10,
                                  textColor=colors.grey)
    value_style = ParagraphStyle('Value', parent=styles['Normal'], fontSize=12,
                                  spaceAfter=15)

    elements = []

    # Titel (mit Belegnummer falls vorhanden)
    if beleg_nr:
        elements.append(Paragraph(f"Bewirtungsbeleg Nr. {beleg_nr}", title_style))
    else:
        elements.append(Paragraph("Bewirtungsbeleg", title_style))
    elements.append(Paragraph("gemäß § 4 Abs. 5 Nr. 2 EStG", subtitle_style))
    elements.append(Spacer(1, 10*mm))

    # Hauptdaten als Tabelle
    main_data = [
        ['Tag der Bewirtung:', datum],
        ['Ort (Name und Anschrift):', f"{restaurant}, {ort}" if ort else restaurant],
    ]

    main_table = Table(main_data, colWidths=[60*mm, 110*mm])
    main_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('LINEBELOW', (1, 0), (1, -1), 0.5, colors.lightgrey),
    ]))
    elements.append(main_table)
    elements.append(Spacer(1, 8*mm))

    # Teilnehmer-Tabelle
    elements.append(Paragraph("Bewirtete Personen:", label_style))
    elements.append(Spacer(1, 3*mm))

    if teilnehmer:
        teilnehmer_data = [['Name', 'Firma/Funktion']]
        for t in teilnehmer:
            if isinstance(t, dict):
                teilnehmer_data.append([t.get('name', ''), t.get('firma', '')])
            else:
                teilnehmer_data.append([str(t), ''])
    else:
        # Leere Zeilen zum Ausfüllen
        teilnehmer_data = [['Name', 'Firma/Funktion']]
        for _ in range(4):
            teilnehmer_data.append(['', ''])

    teilnehmer_table = Table(teilnehmer_data, colWidths=[85*mm, 85*mm])
    teilnehmer_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f5f5f5')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('MINROWHEIGHT', (0, 1), (-1, -1), 10*mm),
    ]))
    elements.append(teilnehmer_table)
    elements.append(Spacer(1, 8*mm))

    # Anlass
    anlass_data = [
        ['Anlass der Bewirtung:', anlass if anlass else ''],
    ]
    anlass_table = Table(anlass_data, colWidths=[60*mm, 110*mm])
    anlass_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('LINEBELOW', (1, 0), (1, -1), 0.5, colors.lightgrey),
        ('MINROWHEIGHT', (0, 0), (-1, -1), 12*mm),
    ]))
    elements.append(anlass_table)
    elements.append(Spacer(1, 8*mm))

    # Betrag
    betrag_data = [
        ['Höhe der Aufwendungen:', f"{betrag:.2f} €" if betrag else '______________ €'],
    ]
    betrag_table = Table(betrag_data, colWidths=[60*mm, 110*mm])
    betrag_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('FONTSIZE', (1, 0), (1, -1), 14),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(betrag_table)
    elements.append(Spacer(1, 15*mm))

    # Unterschrift
    # Prüfen ob eine Unterschrift vorhanden ist
    sig_image = None
    if unterschrift_base64:
        try:
            # Base64-Daten dekodieren (entferne data:image/png;base64, Prefix falls vorhanden)
            if ',' in unterschrift_base64:
                unterschrift_base64 = unterschrift_base64.split(',')[1]
            img_data = base64.b64decode(unterschrift_base64)
            sig_buffer = io.BytesIO(img_data)
            sig_image = RLImage(sig_buffer, width=37.5*mm, height=15*mm)  # 75% der Originalgröße
        except Exception as e:
            print(f"Fehler beim Laden der Unterschrift: {e}")
            sig_image = None

    if sig_image:
        # Mit Unterschriftsbild
        sig_data = [
            [sig_image, ''],
            ['Datum, Unterschrift des Bewirtenden', bewirtende_person],
        ]
    else:
        # Ohne Unterschrift - leeres Feld zum Unterschreiben
        sig_data = [
            ['', ''],
            ['Datum, Unterschrift des Bewirtenden', bewirtende_person],
        ]

    sig_table = Table(sig_data, colWidths=[85*mm, 85*mm])
    sig_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (0, 1), (0, 1), colors.grey),
        ('FONTSIZE', (1, 1), (1, 1), 11),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'BOTTOM'),
        ('LINEABOVE', (0, 1), (0, 1), 0.5, colors.black),
        ('LINEABOVE', (1, 1), (1, 1), 0.5, colors.black),
        ('TOPPADDING', (0, 1), (-1, 1), 5),
        ('MINROWHEIGHT', (0, 0), (-1, 0), 15*mm),
    ]))
    elements.append(sig_table)

    # Hinweis
    elements.append(Spacer(1, 15*mm))
    hinweis_style = ParagraphStyle('Hinweis', parent=styles['Normal'], fontSize=8,
                                    textColor=colors.grey, alignment=1)
    elements.append(Paragraph(
        "Hinweis: Bitte Originalbeleg anheften. Bei Bewirtungen in Gaststätten ist dieser Beleg "
        "zusammen mit der Rechnung der Gaststätte aufzubewahren.",
        hinweis_style
    ))

    doc.build(elements)
    output.seek(0)

    # Optimiertes Namensschema: YYYY-MM-DD_Restaurant_Bewirtungsbeleg.pdf
    # Datum von DD.MM.YYYY zu YYYY-MM-DD konvertieren
    try:
        datum_parts = datum.split('.')
        if len(datum_parts) == 3:
            iso_datum = f"{datum_parts[2]}-{datum_parts[1]}-{datum_parts[0]}"
        else:
            iso_datum = datum.replace('.', '-')
    except Exception:
        iso_datum = datum.replace('.', '-')

    # Restaurant-Name säubern
    safe_restaurant = re.sub(r'[^\w\s-]', '', restaurant).strip()
    safe_restaurant = re.sub(r'\s+', '_', safe_restaurant)[:25]

    # Dateiname mit Belegnummer falls vorhanden
    if beleg_nr:
        filename = f"{beleg_nr:02d}_{iso_datum}_{safe_restaurant}_Bewirtungsbeleg.pdf"
    else:
        filename = f"{iso_datum}_{safe_restaurant}_Bewirtungsbeleg.pdf"

    # Ordner für Kostenerstattung erstellen (exports/Jahr/Monat/bewirtungsbelege/)
    bewirtungsbelege_dir = get_export_dir(monat, subfolder='bewirtungsbelege')

    # PDF in Ordner speichern
    filepath = os.path.join(bewirtungsbelege_dir, filename)
    with open(filepath, 'wb') as f:
        f.write(output.getvalue())

    output.seek(0)
    return send_file(output, mimetype='application/pdf', as_attachment=True, download_name=filename)

@app.route('/export/zip', methods=['POST'])
def export_zip():
    """Exportiert alles in ein ZIP: Excel, PDF und alle Belege."""
    import zipfile

    data = request.json
    meta = data.get('meta', {})
    expenses = data.get('expenses', {})
    monat = meta.get('monat', 'Spesen')

    # ZIP-Buffer erstellen
    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        # 1. Excel exportieren
        excel_buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = monat

        # Styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='333333')
        title_font = Font(color='333333', size=14, bold=True)

        # Header
        ws['A1'] = monat
        ws['A1'].font = title_font
        ws['C1'] = meta.get('name', '')
        ws['C1'].font = title_font
        ws['F1'] = f"Datum {meta.get('datum', datetime.now().strftime('%d.%m.%y'))}"

        row = 3
        gesamt = 0

        for cat_key, cat_info in CATEGORIES.items():
            cat_expenses = expenses.get(cat_key, [])
            cat_expenses = sort_expenses_by_date(cat_expenses)
            cat_sum = 0

            ws.cell(row=row, column=1, value=f"{list(CATEGORIES.keys()).index(cat_key) + 1}. {cat_info['name']}")
            ws.cell(row=row, column=1).fill = header_fill
            ws.cell(row=row, column=1).font = header_font

            if cat_key == 'fahrtkosten_kfz':
                headers = ['', 'Datum', 'Fahrstrecke', 'Anlaß', 'km', '0,3', 'Betrag €']
            else:
                headers = ['', 'Datum', 'Beschreibung', '', '', '', 'Betrag €']

            for i, h in enumerate(headers, 2):
                ws.cell(row=row, column=i, value=h)
                ws.cell(row=row, column=i).fill = header_fill
                ws.cell(row=row, column=i).font = header_font
            row += 1

            for exp in cat_expenses:
                if cat_key == 'fahrtkosten_kfz':
                    km = float(exp.get('km', 0) or 0)
                    betrag = km * 0.30
                    ws.cell(row=row, column=2, value=exp.get('datum', ''))
                    ws.cell(row=row, column=3, value=exp.get('fahrstrecke', ''))
                    ws.cell(row=row, column=4, value=exp.get('anlass', ''))
                    ws.cell(row=row, column=5, value=km)
                    ws.cell(row=row, column=6, value='0,30')
                    ws.cell(row=row, column=7, value=f"{betrag:.2f}")
                else:
                    betrag = float(exp.get('betrag', 0) or 0)
                    ws.cell(row=row, column=2, value=exp.get('datum', exp.get('monat', '')))
                    beschreibung = exp.get('beschreibung', exp.get('personen', exp.get('ort', '')))
                    ws.cell(row=row, column=3, value=beschreibung)
                    ws.cell(row=row, column=7, value=f"{betrag:.2f}")
                cat_sum += betrag
                row += 1

            ws.cell(row=row, column=6, value="Summe:")
            ws.cell(row=row, column=6).font = Font(bold=True)
            ws.cell(row=row, column=7, value=f"{cat_sum:.2f}")
            ws.cell(row=row, column=7).font = Font(bold=True)
            gesamt += cat_sum
            row += 2

        ws.cell(row=row, column=6, value="GESAMT:")
        ws.cell(row=row, column=6).font = Font(size=12, bold=True)
        ws.cell(row=row, column=7, value=f"{gesamt:.2f}")
        ws.cell(row=row, column=7).font = Font(size=12, bold=True)

        wb.save(excel_buffer)
        excel_buffer.seek(0)
        zf.writestr(f"Spesen_{monat.replace(' ', '_')}.xlsx", excel_buffer.getvalue())

        # 2. PDF-Übersicht hinzufügen
        pdf_buffer = generate_pdf_buffer(meta, expenses)
        zf.writestr(f"Spesen_{monat.replace(' ', '_')}.pdf", pdf_buffer.getvalue())

        # 3. Original-Belege sammeln (per file_hash) mit fortlaufender Nummerierung
        cache = {}
        if os.path.exists(CACHE_FILE):
            with open(CACHE_FILE, 'r') as f:
                cache = json.load(f)

        beleg_nr = 1  # Fortlaufende Belegnummer für Dateinamen
        for cat_key in CATEGORIES.keys():
            cat_expenses = expenses.get(cat_key, [])
            # Nach Datum sortieren (gleiche Reihenfolge wie in Excel/PDF)
            cat_expenses = sort_expenses_by_date(cat_expenses)

            for exp in cat_expenses:
                file_hash = exp.get('file_hash')
                if file_hash and file_hash in cache:
                    cache_entry = cache[file_hash]
                    datei_pfad = cache_entry.get('datei_pfad')
                    datei_name = cache_entry.get('datei', f'beleg_{beleg_nr}.pdf')

                    if datei_pfad and os.path.exists(datei_pfad):
                        # Dateiendung extrahieren
                        _, ext = os.path.splitext(datei_name)
                        # Beleg mit Nummer-Präfix ins ZIP hinzufügen
                        numbered_name = f"{beleg_nr:02d}_{datei_name}"
                        zf.write(datei_pfad, f"Belege/{numbered_name}")

                beleg_nr += 1  # Nummer erhöhen für jeden Eintrag (auch ohne Beleg)

        # 4. Bewirtungsbelege aus dem Export-Ordner hinzufügen
        bewirtungsbelege_dir = get_export_dir(monat, subfolder='bewirtungsbelege')
        if os.path.exists(bewirtungsbelege_dir):
            for filename in os.listdir(bewirtungsbelege_dir):
                if filename.endswith('.pdf'):
                    filepath = os.path.join(bewirtungsbelege_dir, filename)
                    zf.write(filepath, f"Bewirtungsbelege/{filename}")

    zip_buffer.seek(0)

    # Dateiname für ZIP
    safe_monat = monat.replace(' ', '_').replace('/', '-')
    zip_filename = f"Spesen_{safe_monat}_komplett.zip"

    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=zip_filename
    )


if __name__ == '__main__':
    app.run(debug=True, port=5000)
